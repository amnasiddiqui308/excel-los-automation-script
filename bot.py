from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from math import *

wb = load_workbook("Example0_LOS Original.xlsx")
wb2 = load_workbook("Example0gross_LOSDesignation.xlsx")

ws = wb.active
ws2 = wb2.active

names = list(dict.fromkeys([cell.value for cell in ws['A']], None).keys())
names.remove("Account")
print(names)

description = list(cell.value for cell in ws2['A'])
description.remove("Description")
#print(description)

wb_out = Workbook()
ws_out = wb_out.active

ws_row = 2
ws_out_row = 2
for i in range(1, ws.max_column + 1):
    ws_out.cell(1, i).value = ws.cell(1, i).value
    ws_out.cell(1, i).number_format = ws.cell(1, i).number_format
    ws_out.cell(1, i).font = Font(bold=True)   

for i in names:
    for j in description:
        if i == ws.cell(ws_row, 1).value and j == ws.cell(ws_row, 2).value:
            for k in range(1, ws.max_column + 1):

                if not isinstance(ws.cell(ws_row, k).value, (float, int)):
                    ws_out.cell(ws_out_row, k).value = ws.cell(ws_row, k).value 
                else:
                    ws_out.cell(ws_out_row, k).value = ws.cell(ws_row, k).value
                    ws_out.cell(ws_out_row, k).number_format = "#,##0.00"

            ws_row += 1

        else:
            ws_out.cell(ws_out_row, 1).value = i
            ws_out.cell(ws_out_row, 2).value = j
            for k in range(3, ws.max_column + 1):
                ws_out.cell(ws_out_row, k).value = 0
                ws_out.cell(ws_out_row, k).number_format = "#,##0.00"

        ws_out_row += 1

ws_out.freeze_panes = ws_out['A2']
wb_out.save("updated_sheet.xlsx")